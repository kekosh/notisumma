# -*- coding: utf-8 -*-

import openpyxl
import datetime as dt

'''
Excelファイルのデータを読み込む
memo:作成中
'''



'''
テキスデータをExcelファイルに出力する
memo:
'''
def create_xlbook(list_read_data,base_xlfile):
    wb = openpyxl.load_workbook(base_xlfile)
    ws = wb.worksheets[0]
    
    # 「内容」列にデータが存在する最終行を取得
    _row = 5
    while True:
        if ws.cell(_row,7).value is None:
            break
        _row += 1
    for data in list_read_data:
        ws.cell(row=_row,column=1,value=_row)
        ws.cell(row=_row,column=2,value='{0:%Y/%m/%d}'.format(dt.date.today()))
        ws.cell(row=_row,column=3,value=data['title'])
        ws.cell(row=_row,column=4,value=data['product'])
        ws.cell(row=_row,column=5,value=data['category'])
        ws.cell(row=_row,column=6,value=data['version'])
        
        '''
        1行のセルの高さ最大設定時に表示できる行数をLINES_LIMITに設定
        (デフォルト：MS Pゴシック 11pt の場合、30行まで)
        '''
        LINES_LIMIT = 30
        lines = data['description'].split('\n')
        _lines_all = len(lines)
        # 内容列セル出力分割処理
        if _lines_all > LINES_LIMIT :
            _line_cnt, _last_line_cnt = 0, 0
            while True:
                # 行数上限単位に分割して出力
                if _line_cnt % LINES_LIMIT == 0 and _line_cnt != 0:
                    ws.cell(row=_row,column=7,value='\n'.join(lines[_last_line_cnt:_line_cnt]).strip())
                    _last_line_cnt = _line_cnt
                    _row += 1 
                    _line_cnt += 1
                    continue
                _line_cnt += 1
                # 残り行数が上限に満たない場合
                if _lines_all % LINES_LIMIT == _lines_all - _line_cnt:
                    ws.cell(row=_row,column=7,value='\n'.join(lines[_line_cnt:]).strip())
                    _row += 1
                    break
                if _line_cnt > _lines_all:
                    break
        else:
            ws.cell(row=_row,column=7,value=data['description'])
            _row += 1
    try:
        wb.save(base_xlfile)
    except PermissionError:
        raise PermissionError('ファイルにアクセスできませんでした。')
    except Exception:
        raise Exception('不明なエラーが発生しました。')

if __name__ == '__main__':
    xlfile = r"D:\MyFile\arc\venv\notisumma\extract\test.xlsx"
    readxl(xlfile)
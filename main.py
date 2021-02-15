import glob
import os
import zipfile
from openpyxl import load_workbook


# ファイル展開処理
def unzip(filepath, extract_path, password):
    with zipfile.ZipFile(filepath, 'r') as zip_file:
        for fileinfo in zip_file.infolist():
            fileinfo.filename = fileinfo.filename.encode('cp437').decode('cp932')
            zip_file.extract(fileinfo ,path=extract_path, pwd=password.encode())

# Excelタイプ
def load_excel_data(xlfile, sheetname):
    wb = load_workbook(filename = xlfile)
    xlsheet = wb[sheetname]
    d = xlsheet.max_column
    print(d)

# テキストタイプ
def read_txt(path, filename):

    KEY_OVERVIEW = '＜現象内容＞'
    KEY_CAUSE_RESPONSE = '＜原因及び対応策＞'
    file_path = os.path.join(path,filename)


    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            _first_line = file.readline()

            # 管理番号取得
            _manage_number_posi = _first_line.find('＞')
            _manage_number = _first_line[:_manage_number_posi].strip('＜')

            # 通知種別取得
            _notice_type_posi_start = _first_line.find('【') + 1
            _notice_type_posi_end = _first_line.find('】')
            _notice_type = _first_line[_notice_type_posi_start:_notice_type_posi_end]

            # 現象内容取得
            _overview_start = 0
            _overview_end = 0
            _overview = ''
            # for line in file.readlines():
            #     if line in KEY_OVERVIEW:
            #         _overview_start = line.find(KEY_OVERVIEW)
            #     else:
            #         return 'irregular format text.overview keyword is not found'

            #     if line in KEY_CAUSE_RESPONSE:
            #         _overview_end = line.find(KEY_CAUSE_RESPONSE)
            #     else:
            #         return 'irregular format text.cause and response  keyword is not found'
            print('*＜現象内容＞*' in file.readlines())


    else:
        return 'file not found.'



if __name__ == '__main__':
    _base_path = r'D:\work\venv\notisumma\sample'
    _extract_path = r'.\\ext'
    _password = 'password'

    # 拡張子が"zip"であるファイル一覧を取得
    zip_path_list = glob.glob(f'{_base_path}\**\*.zip',recursive=True)

    for f_path in zip_path_list:
        unzip(f_path, _extract_path, _password)

    # # Excelデータ読み取り
    # _xlfile = r"D:\work\venv\notisumma\ext\sample\不具合情報一覧.xlsx"
    # _sheetname = '不具合情報一覧'
    # load_excel_data(_xlfile, _sheetname)

    # # テキストファイル読み取り
    # _dir = r'D:\work\venv\notisumma\ext\sample_txt'
    # filename = '不具合情報.txt'
    # r = read_txt(_dir,filename)
    # print(r)